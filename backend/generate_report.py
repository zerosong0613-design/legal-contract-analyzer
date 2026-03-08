"""
ContractLens — Excel 보고서 생성기
사용법: python3 generate_report.py '<records_json>' <output_path>
"""
import sys
import json
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
try:
    import holidays as _holidays
    def _kr_holidays(year):
        return _holidays.KR(years=year)
except ImportError:
    def _kr_holidays(year):
        return {}
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

# ── 색상 ────────────────────────────────────────────────────
C = {
    "navy":     "1E3A5F",
    "indigo":   "4F46E5",
    "indigo_lt":"EEF2FF",
    "slate":    "1E293B",
    "slate_lt": "F8FAFC",
    "white":    "FFFFFF",
    "green":    "16A34A",  "green_lt":  "DCFCE7",
    "amber":    "D97706",  "amber_lt":  "FEF9C3",
    "red":      "DC2626",  "red_lt":    "FEE2E2",
    "crimson":  "991B1B",
    "gray":     "64748B",
    "border":   "CBD5E1",
}
GRADE_STYLE = {
    "L1":("16A34A","DCFCE7"), "L2":("D97706","FEF9C3"), "L3":("DC2626","FEE2E2"),
    "R1":("16A34A","DCFCE7"), "R2":("D97706","FEF9C3"), "R3":("DC2626","FEE2E2"),
}

# ── 헬퍼 ────────────────────────────────────────────────────
def F(bold=False, color="1E293B", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def BG(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def working_days(d0, d1):
    """d0(수신일)~d1(회신일) 사이 영업일수 (주말·한국공휴일 제외, 당일 미포함 익일부터 카운트)"""
    from datetime import timedelta
    if d1 < d0:
        return 0
    years = set(range(d0.year, d1.year + 1))
    hols = {}
    for y in years:
        hols.update(_kr_holidays(y))
    count = 0
    cur = d0 + timedelta(days=1)
    while cur <= d1:
        if cur.weekday() < 5 and cur not in hols:
            count += 1
        cur += timedelta(days=1)
    return count

def AL(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def BD(color="CBD5E1", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, row, text, s, e, bg=None, fg="FFFFFF", size=14, height=34):
    bg = bg or C["navy"]
    ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    c = ws.cell(row=row, column=s, value=text)
    c.font = F(bold=True, color=fg, size=size)
    c.fill = BG(bg)
    c.alignment = AL("center", "center")
    ws.row_dimensions[row].height = height

def sub_row(ws, row, text, s, e, bg="EEF2FF", fg="64748B", height=18):
    ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    c = ws.cell(row=row, column=s, value=text)
    c.font = F(color=fg, size=9, italic=True)
    c.fill = BG(bg)
    c.alignment = AL("center", "center")
    ws.row_dimensions[row].height = height

def header_row(ws, row, headers, start_col=1, bg="1E3A5F", fg="FFFFFF", height=20):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+i, value=h)
        c.font = F(bold=True, color=fg, size=9)
        c.fill = BG(bg)
        c.alignment = AL("center", "center")
        c.border = BD()

def data_cell(ws, row, col, val, bg="FFFFFF", bold=False, align_h="center", size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = F(bold=bold, color=C["slate"], size=size)
    c.fill = BG(bg)
    c.alignment = AL(align_h, "center", wrap)
    c.border = BD()
    return c

def grade_cell(ws, row, col, grade):
    fg, bg = GRADE_STYLE.get(grade, (C["gray"], C["slate_lt"]))
    c = ws.cell(row=row, column=col, value=grade)
    c.font = F(bold=True, color=fg, size=10)
    c.fill = BG(bg)
    c.alignment = AL("center", "center")
    c.border = BD()

def section_title(ws, row, text, s, e, bg="4F46E5", fg="FFFFFF", height=22):
    ws.merge_cells(start_row=row, start_column=s, end_row=row, end_column=e)
    c = ws.cell(row=row, column=s, value=text)
    c.font = F(bold=True, color=fg, size=11)
    c.fill = BG(bg)
    c.alignment = AL("left", "center")
    ws.row_dimensions[row].height = height

def kpi_card(ws, label_row, val_row, col, label, value, color):
    lc = ws.cell(row=label_row, column=col, value=label)
    lc.font = F(color=C["gray"], size=9)
    lc.fill = BG("F1F5F9")
    lc.alignment = AL("center", "center")
    lc.border = BD()
    ws.row_dimensions[label_row].height = 16
    vc = ws.cell(row=val_row, column=col, value=value)
    vc.font = Font(name="Arial", bold=True, color=color, size=20)
    vc.fill = BG(C["white"])
    vc.alignment = AL("center", "center")
    vc.border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thick", color=color),
    )
    ws.row_dimensions[val_row].height = 38

def get_tags(records):
    tags = defaultdict(int)
    for r in records:
        for t in [r.get("risk_type_1",""), r.get("risk_type_2",""), r.get("risk_type_3","")]:
            if t: tags[t] += 1
    return sorted(tags.items(), key=lambda x: -x[1])

def get_assignees(records):
    d = defaultdict(lambda: {"total":0,"r3":0,"amount":0})
    for r in records:
        a = r.get("assignee") or "미지정"
        d[a]["total"]  += 1
        d[a]["r3"]     += 1 if r.get("risk_grade")=="R3" else 0
        d[a]["amount"] += r.get("amount",0) or 0
    return d

def quarter_of(ym):
    m = int(ym[5:7])
    return (m-1)//3 + 1

COLORS_BAR = [C["indigo"],"059669","D97706","DC2626","7C3AED","0891B2"]

# ════════════════════════════════════════════════════════════
def generate(records, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    today = datetime.today().strftime("%Y년 %m월 %d일")
    total = len(records)

    # ── 집계 ────────────────────────────────────────────────
    monthly_agg  = defaultdict(lambda:{"total":0,"l1":0,"l2":0,"l3":0,"r1":0,"r2":0,"r3":0,"amount":0})
    quarterly    = defaultdict(lambda:{"total":0,"l1":0,"l2":0,"l3":0,"r1":0,"r2":0,"r3":0,"amount":0})
    yearly       = defaultdict(lambda:{"total":0,"l1":0,"l2":0,"l3":0,"r1":0,"r2":0,"r3":0,"amount":0})
    for r in records:
        d = r.get("date","")
        if not d: continue
        m = d[:7]
        q = f"{d[:4]} Q{quarter_of(d[:7])}"
        y = d[:4]
        for agg, key in [(monthly_agg, m),(quarterly, q),(yearly, y)]:
            agg[key]["total"]  += 1
            agg[key][r.get("lead_grade","").lower()] += 1
            agg[key][r.get("risk_grade","").lower()] += 1
            agg[key]["amount"] += r.get("amount",0) or 0

    all_tags   = get_tags(records)
    total_tags = sum(v for _,v in all_tags)
    assignees  = get_assignees(records)

    # ══════════════════════════════════════════════════════
    # ① 기본 로그
    # ══════════════════════════════════════════════════════
    ws1 = wb.create_sheet("① 기본로그")
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A4"
    set_widths(ws1, [12,28,18,18,16,8,8,18,12,10,8,10,14,12,20,15,15,15,10])

    LEAD_TARGET = {"L1": 2, "L2": 5, "L3": None}


    title_row(ws1, 1, "📋  계약 기본 로그 — 전 계약 대상", 1, 20, C["navy"])
    sub_row(ws1, 2, f"생성일: {today}  |  총 {total}건  |  R3 반드시 태깅·R2 선택·R1 생략 가능", 1, 20)
    ws1.row_dimensions[3].height = 4

    HEADERS1 = ["수신일","계약명","우리측 당사자","거래상대방","계약유형","등급(L)","등급(R)","리스크 유형",
                "계약금액\n(백만원)","회신일","소요\n일수","목표\n일수","달성\n여부","구조변화\n(R3만)",
                "담당자","Risk_Type_1\n(주요)","Risk_Type_2\n(부수)","Risk_Type_3\n(기타)","비고","태깅완료"]
    header_row(ws1, 4, HEADERS1, height=32)

    for i, r in enumerate(records):
        row = 5 + i
        ws1.row_dimensions[row].height = 18
        bg  = C["white"] if i%2==0 else C["slate_lt"]
        reply_date  = r.get("reply_date","")
        recv_date   = r.get("date","")
        lead_g      = r.get("lead_grade","")
        target_days = LEAD_TARGET.get(lead_g)
        # 소요 영업일수 계산 (주말·한국공휴일 제외)
        elapsed = ""
        achieved = ""
        if reply_date and recv_date:
            try:
                from datetime import date as _d
                d0 = _d.fromisoformat(recv_date)
                d1 = _d.fromisoformat(reply_date)
                elapsed = working_days(d0, d1)
                if target_days is not None:
                    achieved = "✅ 달성" if elapsed <= target_days else f"❌ {elapsed-target_days}일 초과"
                else:
                    achieved = "협의"
            except:
                pass

        vals = [recv_date, r.get("title",""),
                r.get("our_party","") or "SK케미칼",
                r.get("counterparty",""), r.get("contract_type",""),
                None, None,
                r.get("risk_type_1","") or "-",
                r.get("amount","") or "-",
                reply_date or "-",
                elapsed if elapsed != "" else "-",
                target_days if target_days else "협의",
                achieved or "-",
                r.get("note","") or "-",
                r.get("assignee","") or "-",
                r.get("risk_type_1","") or "-",
                r.get("risk_type_2","") or "-",
                r.get("risk_type_3","") or "-",
                r.get("memo","") or "-",
                "Y"]
        for col, val in enumerate(vals, 1):
            if val is None: continue
            al = "center" if col in [1,6,7,9,10,11,12,13,20] else "left"
            c = data_cell(ws1, row, col, val, bg, align_h=al, size=9)
            # 달성여부 색상
            if col == 13 and achieved:
                if "달성" in str(achieved):
                    c.font = F(bold=True, color=C["green"], size=9)
                elif "초과" in str(achieved):
                    c.font = F(bold=True, color=C["red"], size=9)
        grade_cell(ws1, row, 6, r.get("lead_grade",""))
        grade_cell(ws1, row, 7, r.get("risk_grade",""))

    ws1.sheet_properties.tabColor = C["navy"]

    # ══════════════════════════════════════════════════════
    # ② R3 가치기록
    # ══════════════════════════════════════════════════════
    ws2 = wb.create_sheet("② R3 가치기록")
    ws2.sheet_view.showGridLines = False
    set_widths(ws2, [12,28,14,18,32,32,32,18,10,12,15,15,15])

    title_row(ws2, 1, "⭐  R3 고위험 계약 가치 기록 — 선별 20%, 임원 보고·성과 증빙", 1, 13, C["crimson"])
    sub_row(ws2, 2, "목적: 임원 보고·성과 증빙  |  3줄 고정 포맷 + 개선수준(1~3) + 보고대상(Y/N)", 1, 13, "FFF5F5","991B1B")
    ws2.row_dimensions[3].height = 6

    HEADERS2 = ["날짜","계약명","금액(백만원)","거래상대방","① 초기 구조",
                "② 법무 제안 대안","③ 최종 합의+잔여리스크",
                "개선수준\n1방어/2구조/3전략","보고대상\nY/N","담당자",
                "Risk_Type_1","Risk_Type_2","Risk_Type_3"]
    header_row(ws2, 4, HEADERS2, bg=C["crimson"], height=32)

    r3s = [r for r in records if r.get("risk_grade")=="R3"]
    for i, r in enumerate(r3s):
        row = 5 + i
        ws2.row_dimensions[row].height = 40
        bg  = "FFF5F5" if i%2==0 else "FEF2F2"
        vals = [r.get("date",""), r.get("title",""), r.get("amount","") or "-",
                r.get("counterparty",""),
                "(초기 계약 구조 입력)", "(법무팀 협상 제안)", "(최종 합의 결과)",
                "1_방어", "Y", r.get("assignee",""),
                r.get("risk_type_1","") or "-",
                r.get("risk_type_2","") or "-",
                r.get("risk_type_3","") or "-"]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = F(color=C["slate"], size=9)
            c.fill = BG(bg)
            c.alignment = AL("left" if col in [5,6,7] else "center", "center", wrap=col in [5,6,7])
            c.border = BD()

    ws2.sheet_properties.tabColor = C["crimson"]

    # ══════════════════════════════════════════════════════
    # ③ 월별 성과분석
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("③ 월별 성과분석")
    ws3.sheet_view.showGridLines = False
    set_widths(ws3, [12,8,8,8,8,8,8,8,14,22,10,10,10])

    title_row(ws3, 1, "📈  월별 성과 분석", 1, 13, "1D4ED8")
    sub_row(ws3, 2, f"생성일: {today}  |  월별 계약 건수·등급 분포·금액 집계", 1, 13, "EFF6FF","1D4ED8")
    ws3.row_dimensions[3].height = 8

    section_title(ws3, 4, "📊  누계 KPI", 1, 9, "1D4ED8")
    LEAD_TARGET = {"L1": 2, "L2": 5, "L3": None}
    # 달성률 계산 (회신일 있는 건만)
    replied = [r for r in records if r.get("reply_date") and r.get("date")]
    achieved_cnt = 0
    for r in replied:
        try:
            from datetime import date as _d
            d0 = _d.fromisoformat(r["date"])
            d1 = _d.fromisoformat(r["reply_date"])
            td = LEAD_TARGET.get(r.get("lead_grade",""))
            wd = working_days(d0, d1)
            if td is not None and wd <= td:
                achieved_cnt += 1
            elif td is None:
                achieved_cnt += 1  # L3 협의건은 달성으로 집계
        except: pass
    achieve_rate = f"{achieved_cnt/len(replied)*100:.0f}%" if replied else "-"

    kpi_items = [
        ("전체", total, C["indigo"]),
        ("L1", sum(1 for r in records if r.get("lead_grade")=="L1"), C["green"]),
        ("L2", sum(1 for r in records if r.get("lead_grade")=="L2"), C["amber"]),
        ("L3", sum(1 for r in records if r.get("lead_grade")=="L3"), C["red"]),
        ("R1", sum(1 for r in records if r.get("risk_grade")=="R1"), C["green"]),
        ("R2", sum(1 for r in records if r.get("risk_grade")=="R2"), C["amber"]),
        ("R3", sum(1 for r in records if r.get("risk_grade")=="R3"), C["red"]),
        ("금액(백만)", sum(r.get("amount",0) or 0 for r in records), "0F172A"),
        ("리드타임\n달성률", achieve_rate, C["green"]),
    ]
    for i, (lbl, val, color) in enumerate(kpi_items):
        kpi_card(ws3, 5, 6, i+1, lbl, val, color)
    ws3.row_dimensions[7].height = 12

    section_title(ws3, 8, "📅  월별 상세 현황", 1, 9, "1D4ED8")
    header_row(ws3, 9, ["월","전체","L1","L2","L3","R1","R2","R3","금액(백만)"], bg="1D4ED8")
    sorted_months = sorted(monthly_agg.keys())
    for i, m in enumerate(sorted_months):
        v   = monthly_agg[m]
        row = 10 + i
        ws3.row_dimensions[row].height = 18
        bg  = C["white"] if i%2==0 else C["slate_lt"]
        for col, val in enumerate([m,v["total"],v["l1"],v["l2"],v["l3"],v["r1"],v["r2"],v["r3"],v["amount"]], 1):
            c = data_cell(ws3, row, col, val, bg, align_h="center")
            if col==8 and val>0: c.font = F(bold=True, color=C["red"])

    if sorted_months:
        bar = BarChart()
        bar.type = "col"
        bar.title = "월별 계약 건수"
        bar.style = 10
        bar.width = 18
        bar.height = 12
        bar.add_data(Reference(ws3, min_col=2, max_col=8, min_row=9, max_row=9+len(sorted_months)), titles_from_data=True)
        bar.set_categories(Reference(ws3, min_col=1, min_row=10, max_row=9+len(sorted_months)))
        ws3.add_chart(bar, f"A{12+len(sorted_months)}")

    section_title(ws3, 9, "👤  담당자별 현황", 10, 13, "1D4ED8")
    header_row(ws3, 10, ["담당자","전체","R3건수","금액(백만)"], start_col=10, bg="1D4ED8")
    for i, (a, v) in enumerate(sorted(assignees.items())):
        row = 11 + i
        ws3.row_dimensions[row].height = 18
        bg = C["white"] if i%2==0 else C["slate_lt"]
        for col, val in enumerate([a, v["total"], v["r3"], v["amount"]], 10):
            data_cell(ws3, row, col, val, bg, align_h="left" if col==10 else "center")

    ws3.sheet_properties.tabColor = "1D4ED8"

    # ══════════════════════════════════════════════════════
    # ④ 분기 성과분석
    # ══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("④ 분기 성과분석")
    ws4.sheet_view.showGridLines = False
    set_widths(ws4, [14,8,8,8,8,8,8,8,16,22,10,10,10])

    title_row(ws4, 1, "📊  분기 성과 분석", 1, 13, "065F46")
    sub_row(ws4, 2, f"생성일: {today}  |  Q1~Q4 분기별 집계  |  리스크 태그 분포 포함", 1, 13, "ECFDF5","065F46")
    ws4.row_dimensions[3].height = 8

    section_title(ws4, 4, "📅  분기별 상세 현황", 1, 8, "065F46")
    header_row(ws4, 5, ["분기","전체","L1","L2","L3","R1","R2","R3","금액(백만)"], bg="065F46")
    sorted_quarters = sorted(quarterly.keys())
    for i, q in enumerate(sorted_quarters):
        v   = quarterly[q]
        row = 6 + i
        ws4.row_dimensions[row].height = 20
        bg  = C["white"] if i%2==0 else C["slate_lt"]
        for col, val in enumerate([q,v["total"],v["l1"],v["l2"],v["l3"],v["r1"],v["r2"],v["r3"],v["amount"]], 1):
            c = data_cell(ws4, row, col, val, bg, align_h="center")
            if col==8 and val>0: c.font = F(bold=True, color=C["red"])

    base_k = 6 + len(sorted_quarters) + 2
    section_title(ws4, base_k, "📊  분기 KPI 카드", 1, 8, "065F46")
    for i, q in enumerate(sorted_quarters):
        v     = quarterly[q]
        row_l = base_k + 1 + i * 5
        ws4.merge_cells(start_row=row_l, start_column=1, end_row=row_l, end_column=8)
        hc = ws4.cell(row=row_l, column=1, value=f"◆  {q}")
        hc.font = F(bold=True, color="FFFFFF", size=10)
        hc.fill = BG("047857")
        hc.alignment = AL("left","center")
        ws4.row_dimensions[row_l].height = 18
        for j, (lbl, val, color) in enumerate([
            ("전체",v["total"],C["indigo"]),("L1",v["l1"],C["green"]),
            ("L2",v["l2"],C["amber"]),("L3",v["l3"],C["red"]),
            ("R1",v["r1"],C["green"]),("R2",v["r2"],C["amber"]),
            ("R3",v["r3"],C["red"]),("금액",v["amount"],"0F172A"),
        ]):
            kpi_card(ws4, row_l+1, row_l+2, j+1, lbl, val, color)
        ws4.row_dimensions[row_l+4].height = 10

    section_title(ws4, 4, "🏷  리스크 태그 현황", 10, 13, "065F46")
    header_row(ws4, 5, ["리스크 유형","건수","비율","시각화"], start_col=10, bg="065F46")
    for i, (tag, cnt) in enumerate(all_tags):
        row = 6 + i
        ws4.row_dimensions[row].height = 18
        bg  = C["white"] if i%2==0 else C["slate_lt"]
        pct = cnt/total_tags if total_tags else 0
        for col, val in zip([10,11,12],[tag, cnt, f"{pct:.0%}"]):
            data_cell(ws4, row, col, val, bg, align_h="left" if col==10 else "center")
        bc = ws4.cell(row=row, column=13, value="█"*max(1,int(pct*25)))
        bc.font = Font(name="Arial", color=COLORS_BAR[i%len(COLORS_BAR)], size=10)
        bc.fill = BG(bg); bc.alignment=AL("left"); bc.border=BD()

    ws4.sheet_properties.tabColor = "065F46"

    # ══════════════════════════════════════════════════════
    # ⑤ 연간 성과분석
    # ══════════════════════════════════════════════════════
    ws5 = wb.create_sheet("⑤ 연간 성과분석")
    ws5.sheet_view.showGridLines = False
    set_widths(ws5, [16,10,10,10,10,10,10,10,16,22,12,12,12])

    title_row(ws5, 1, "🏆  연간 성과 분석", 1, 13, "6B21A8")
    sub_row(ws5, 2, f"생성일: {today}  |  연도별 집계  |  거래상대방·리스크 연간 트렌드", 1, 13, "FAF5FF","6B21A8")
    ws5.row_dimensions[3].height = 8

    section_title(ws5, 4, "📅  연도별 집계", 1, 8, "6B21A8")
    header_row(ws5, 5, ["연도","전체","L1","L2","L3","R1","R2","R3","금액(백만)"], bg="6B21A8")
    for i, y in enumerate(sorted(yearly.keys())):
        v   = yearly[y]
        row = 6 + i
        ws5.row_dimensions[row].height = 20
        bg  = C["white"] if i%2==0 else C["slate_lt"]
        for col, val in enumerate([y,v["total"],v["l1"],v["l2"],v["l3"],v["r1"],v["r2"],v["r3"],v["amount"]], 1):
            c = data_cell(ws5, row, col, val, bg, align_h="center")
            if col==8 and val>0: c.font = F(bold=True, color=C["red"])

    base_yk = 6 + len(yearly) + 2
    section_title(ws5, base_yk, "🏅  연간 KPI 핵심 지표", 1, 8, "6B21A8")
    r3_count = sum(1 for r in records if r.get("risk_grade")=="R3")
    r3_ratio = r3_count/total if total else 0
    total_amt = sum(r.get("amount",0) or 0 for r in records)
    kpi_annual = [
        ("총 계약", total, C["indigo"]),
        ("R3 건수", r3_count, C["red"]),
        ("R3 비율", f"{r3_ratio:.0%}", C["red"]),
        ("총금액(백만)", f"{total_amt:,}", "0F172A"),
        ("L3 건수", sum(1 for r in records if r.get("lead_grade")=="L3"), C["amber"]),
        ("담당자수", len(set(r.get("assignee","") for r in records if r.get("assignee"))), C["indigo"]),
    ]
    for i, (lbl, val, color) in enumerate(kpi_annual):
        kpi_card(ws5, base_yk+1, base_yk+2, i+1, lbl, val, color)

    base_cp = base_yk + 6
    section_title(ws5, base_cp, "🏢  거래상대방별 연간 현황", 1, 5, "6B21A8")
    header_row(ws5, base_cp+1, ["거래상대방","전체건수","R3건수","R3비율","총금액(백만)"], bg="6B21A8")
    cp = defaultdict(lambda:{"total":0,"r3":0,"amount":0})
    for r in records:
        name = r.get("counterparty","")
        cp[name]["total"]  += 1
        cp[name]["r3"]     += 1 if r.get("risk_grade")=="R3" else 0
        cp[name]["amount"] += r.get("amount",0) or 0
    for i, (name, v) in enumerate(sorted(cp.items(), key=lambda x:-x[1]["total"])):
        row = base_cp + 2 + i
        ws5.row_dimensions[row].height = 18
        bg  = C["white"] if i%2==0 else C["slate_lt"]
        pct = v["r3"]/v["total"] if v["total"] else 0
        for col, val in enumerate([name, v["total"], v["r3"], f"{pct:.0%}", v["amount"]], 1):
            c = data_cell(ws5, row, col, val, bg, align_h="left" if col==1 else "center")
            if col==4 and v["r3"]>0: c.font = F(bold=True, color=C["red"])

    section_title(ws5, 4, "🏷  연간 리스크 태그 Top", 10, 13, "6B21A8")
    header_row(ws5, 5, ["리스크 유형","건수","비율","시각화"], start_col=10, bg="6B21A8")
    for i, (tag, cnt) in enumerate(all_tags):
        row = 6 + i
        ws5.row_dimensions[row].height = 18
        bg  = C["white"] if i%2==0 else C["slate_lt"]
        pct = cnt/total_tags if total_tags else 0
        for col, val in zip([10,11,12],[tag, cnt, f"{pct:.0%}"]):
            data_cell(ws5, row, col, val, bg, align_h="left" if col==10 else "center")
        bc = ws5.cell(row=row, column=13, value="█"*max(1,int(pct*25)))
        bc.font = Font(name="Arial", color=COLORS_BAR[i%len(COLORS_BAR)], size=10)
        bc.fill = BG(bg); bc.alignment=AL("left"); bc.border=BD()

    ws5.sheet_properties.tabColor = "6B21A8"

    # ══════════════════════════════════════════════════════
    # ⑥ 리스크태그 분석
    # ══════════════════════════════════════════════════════
    ws6 = wb.create_sheet("⑥ 리스크태그 분석")
    ws6.sheet_view.showGridLines = False
    set_widths(ws6, [22, 10, 12, 34])

    title_row(ws6, 1, "🏷  리스크 태그 누적 분석", 1, 4, "0F766E")
    sub_row(ws6, 2, "기준: Risk_Type_1·2·3 전체 집계  |  3개월 누적 후 AI 질의 시작 권장", 1, 4, "F0FDFA","0F766E")
    ws6.row_dimensions[3].height = 8

    header_row(ws6, 4, ["리스크 유형","건수","비율(%)","시각화"], bg="0F766E", height=20)
    for i, (tag, cnt) in enumerate(all_tags):
        row = 5 + i
        ws6.row_dimensions[row].height = 20
        bg  = C["white"] if i%2==0 else C["slate_lt"]
        pct = cnt/total_tags if total_tags else 0
        for col, val in zip([1,2,3],[tag, cnt, f"{pct:.1%}"]):
            data_cell(ws6, row, col, val, bg, bold=(col==1), align_h="left" if col==1 else "center")
        bc = ws6.cell(row=row, column=4, value="█"*max(1,int(pct*40)))
        bc.font = Font(name="Arial", color=COLORS_BAR[i%len(COLORS_BAR)], size=11)
        bc.fill = BG(bg); bc.alignment=AL("left"); bc.border=BD()

    if all_tags:
        pie = PieChart()
        pie.title  = "리스크 태그 분포"
        pie.style  = 10
        pie.width  = 14
        pie.height = 12
        pie.add_data(Reference(ws6, min_col=2, min_row=4, max_row=4+len(all_tags)), titles_from_data=True)
        pie.set_categories(Reference(ws6, min_col=1, min_row=5, max_row=4+len(all_tags)))
        ws6.add_chart(pie, f"A{6+len(all_tags)}")

    ws6.sheet_properties.tabColor = "0F766E"

    # 탭 색상
    wb.save(output_path)

# ── 엔트리포인트 ────────────────────────────────────────────
if __name__ == "__main__":
    records_json = sys.argv[1]
    output_path  = sys.argv[2]
    records = json.loads(records_json)
    generate(records, output_path)
    print("ok")
